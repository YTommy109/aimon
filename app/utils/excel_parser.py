"""Excelファイルを解析し、テキストと画像を抽出するユーティリティ。"""

import base64
import io
import json
import logging
from pathlib import Path
from typing import Any, TypedDict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from app.errors import FileReadingError, FileWritingError


class ImageDict(TypedDict):
    figure_number: int
    row: int
    col: int
    marker: str


class SheetDict(TypedDict):
    text_lines: list[str]
    images: list[ImageDict]
    cell_values: dict[str, str]


class JsonDataDict(TypedDict, total=False):
    file_path: str
    total_images: int
    images: list[Any]
    sheets: dict[str, SheetDict]
    text_content: str


class ExcelParser:
    """Excelファイルからテキストと画像を抽出するパーサークラス。"""

    def __init__(self) -> None:
        self.logger = logging.getLogger('aiman')
        self._text_content: str = ''
        self._images: list[Image.Image] = []
        self._parsed_data: dict[str, Any] = {}

    def _process_sheet_images(
        self,
        sheet: Worksheet,
        file_path: Path,
        image_count: int,
    ) -> tuple[dict[tuple[int, int], str], list[dict[str, Any]], list[Image.Image], int]:
        """
        シート内の画像を処理します。

        Args:
            sheet: 処理対象のシート。
            file_path: Excelファイルのパス。
            image_count: 現在の画像カウント。

        Returns:
            画像マーカー辞書、シート画像情報、画像リスト、更新された画像カウントのタプル。
        """
        image_markers: dict[tuple[int, int], str] = {}
        sheet_images: list[dict[str, Any]] = []
        images: list[Image.Image] = []

        if sheet._images:  # type: ignore[attr-defined]
            for img in sheet._images:  # type: ignore[attr-defined]
                # openpyxlの画像オブジェクトからPillowの画像オブジェクトに変換
                try:
                    img_bytes = io.BytesIO(img._data())
                    img_pil = Image.open(img_bytes)
                    images.append(img_pil)

                    # 画像のアンカーから位置を取得
                    row = img.anchor._from.row + 1
                    col = img.anchor._from.col + 1
                    marker = f'[図:{image_count}]'
                    image_markers[(row, col)] = marker

                    # 画像情報を保存
                    sheet_images.append(
                        {'figure_number': image_count, 'row': row, 'col': col, 'marker': marker},
                    )
                    image_count += 1
                except Exception as e:
                    self.logger.warning(f'{file_path}の画像処理に失敗しました: {e}')

        return image_markers, sheet_images, images, image_count

    def _process_sheet_cells(self, sheet: Worksheet) -> dict[tuple[int, int], str]:
        """
        シート内のセルを処理します。

        Args:
            sheet: 処理対象のシート。

        Returns:
            セル値の辞書。
        """
        cell_values: dict[tuple[int, int], str] = {}
        for row_data in sheet.iter_rows():
            for cell in row_data:
                if cell.value is not None:
                    cell_values[(cell.row, cell.column)] = str(cell.value)
        return cell_values

    def _apply_image_markers(
        self,
        cell_values: dict[tuple[int, int], str],
        image_markers: dict[tuple[int, int], str],
    ) -> dict[tuple[int, int], str]:
        """
        セル値に画像マーカーを適用します。

        Args:
            cell_values: セル値の辞書。
            image_markers: 画像マーカーの辞書。

        Returns:
            マーカーが適用されたセル値の辞書。
        """
        for (row, col), marker in image_markers.items():
            if (row, col) in cell_values:
                cell_values[(row, col)] = marker + ' ' + cell_values[(row, col)]
            else:
                cell_values[(row, col)] = marker
        return cell_values

    def _build_sheet_text(self, cell_values: dict[tuple[int, int], str]) -> list[str]:
        """
        セル値からシートのテキストを構築します。

        Args:
            cell_values: セル値の辞書。

        Returns:
            シートのテキスト行のリスト。
        """
        sheet_text_lines: list[str] = []
        if cell_values:
            max_row = max(key[0] for key in cell_values)
            max_col = max(key[1] for key in cell_values)
            for r in range(1, max_row + 1):
                row_text = []
                for c in range(1, max_col + 1):
                    row_text.append(cell_values.get((r, c), ''))
                sheet_text_lines.append('\t'.join(row_text))
        return sheet_text_lines

    def _process_single_sheet(
        self,
        sheet: Worksheet,
        _sheet_name: str,
        file_path: Path,
        image_count: int,
    ) -> tuple[list[str], list[Image.Image], dict[str, Any], int]:
        """
        単一のシートを処理します。

        Args:
            sheet: 処理対象のシート。
            sheet_name: シート名。
            file_path: Excelファイルのパス。
            image_count: 現在の画像カウント。

        Returns:
            テキスト行、画像、シートデータ、更新された画像カウントのタプル。
        """
        # 画像処理
        image_markers, sheet_images, images, updated_count = self._process_sheet_images(
            sheet,
            file_path,
            image_count,
        )

        # セル処理
        cell_values = self._process_sheet_cells(sheet)

        # 画像マーカー適用
        cell_values = self._apply_image_markers(cell_values, image_markers)

        # テキスト構築
        sheet_text_lines = self._build_sheet_text(cell_values)

        # シートデータ構築
        serializable_cell_values = {
            f'{row},{col}': value for (row, col), value in cell_values.items()
        }
        sheet_data = {
            'text_lines': sheet_text_lines,
            'images': sheet_images,
            'cell_values': serializable_cell_values,
        }

        return sheet_text_lines, images, sheet_data, updated_count

    def parse(self, file_path: Path) -> None:
        """
        Excelファイルを解析してテキストと画像を抽出します。

        Args:
            file_path: 処理対象のExcelファイルのパス。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
            FileReadingError: ファイルの読み込みに失敗した場合。
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            raise FileReadingError(str(file_path)) from e

        full_text: list[str] = []
        all_images: list[Image.Image] = []
        image_count = 1
        sheets_data: dict[str, Any] = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f'---シート: {sheet_name}---\n')

            sheet_text_lines, sheet_images, sheet_data, image_count = self._process_single_sheet(
                sheet,
                sheet_name,
                file_path,
                image_count,
            )

            full_text.extend(sheet_text_lines)
            all_images.extend(sheet_images)
            sheets_data[sheet_name] = sheet_data

        self._text_content = '\n'.join(full_text)
        self._images = all_images
        self._parsed_data = {
            'file_path': str(file_path),
            'total_images': len(all_images),
            'sheets': sheets_data,
        }

    def get_text(self) -> str:
        """
        解析されたテキスト内容を取得します。

        Returns:
            抽出されたテキスト内容。
        """
        return self._text_content

    def get_images(self) -> list[Image.Image]:
        """
        解析された画像リストを取得します。

        Returns:
            抽出された画像のリスト。
        """
        return self._images.copy()

    def to_dict(self) -> JsonDataDict:
        """
        解析結果を辞書形式で取得します。

        Returns:
            解析結果の辞書。画像はbase64エンコードされます。
        """
        # 画像をbase64エンコードして追加
        encoded_images: list[Any] = []
        for idx, img in enumerate(self._images):
            buffered = io.BytesIO()
            img.save(buffered, format='PNG')
            img_b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            encoded_images.append({'figure_number': idx + 1, 'format': 'PNG', 'data': img_b64})

        result: JsonDataDict = {
            'file_path': self._parsed_data.get('file_path', ''),
            'total_images': self._parsed_data.get('total_images', 0),
            'sheets': self._parsed_data.get('sheets', {}),
            'images': encoded_images,
            'text_content': self._text_content,
        }
        return result

    def export_json(self, output_path: Path) -> None:
        """
        解析結果をJSONファイルとして出力します。

        Args:
            output_path: 出力するJSONファイルのパス。

        Raises:
            FileWritingError: ファイルの書き込みに失敗した場合。
        """
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.to_dict(), f, ensure_ascii=False, indent=2)
        except Exception as e:
            raise FileWritingError(str(output_path)) from e
