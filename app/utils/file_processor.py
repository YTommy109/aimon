"""ドキュメントからテキストを抽出する最小ユーティリティ。"""

from __future__ import annotations

import logging
from collections.abc import Callable
from pathlib import Path

import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger('aiman')


def read_text(path: Path) -> str:
    """ファイル拡張子に応じてテキストを抽出する。

    Args:
        path: 入力ファイルパス。

    Returns:
        抽出したテキスト。失敗時や非対応形式は空文字。
    """
    suffix = path.suffix.lower()
    text_suffixes = {'.txt', '.md', '.py', '.js', '.html', '.css'}
    handlers: dict[str, Callable[[Path], str]] = {
        '.pdf': _read_pdf_text,
        '.docx': _read_docx_text,
        '.xlsx': _read_xlsx_text,
    }

    if suffix in text_suffixes:
        return _read_plain_text(path)

    handler = handlers.get(suffix)
    return handler(path) if handler else ''


def _read_plain_text(path: Path) -> str:
    try:
        with open(path, encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        logger.warning(f'テキスト読込失敗: {path} ({e})')
        return ''


def _read_pdf_text(path: Path) -> str:
    try:
        text_parts: list[str] = []
        with fitz.open(str(path)) as doc:
            for page in doc:
                text_parts.append(page.get_text())
        return '\n'.join(tp for tp in text_parts if tp)
    except Exception as e:
        logger.warning(f'PDF抽出失敗: {path} ({e})')
        return ''


def _read_docx_text(path: Path) -> str:
    try:
        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text]
        return '\n'.join(paragraphs)
    except Exception as e:
        logger.warning(f'DOCX抽出失敗: {path} ({e})')
        return ''


def _read_xlsx_text(path: Path) -> str:
    try:
        parts = _collect_xlsx_parts(path)
        return '\n'.join(parts)
    except Exception as e:
        logger.warning(f'XLSX抽出失敗: {path} ({e})')
        return ''


def _collect_xlsx_parts(path: Path) -> list[str]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f'# {ws.title}')
        parts.extend(_iter_worksheet_lines(ws))
    wb.close()
    return parts


def _iter_worksheet_lines(ws: Worksheet) -> list[str]:
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None for cell in row):
            continue
        line = ' \t '.join(str(cell) for cell in row if cell is not None)
        if line:
            lines.append(line)
    return lines
