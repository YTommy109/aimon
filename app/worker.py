"""バックグラウンドでプロジェクトを処理するワーカーモジュール。"""

import base64
import io
import json
import logging
import time
from collections.abc import Generator
from contextlib import contextmanager
from multiprocessing import Process
from pathlib import Path
from uuid import UUID

import google.generativeai as genai
import openpyxl
import pandas as pd
from PIL import Image

from app.model import DataManager, Project

from .config import config
from .errors import (
    APICallFailedError,
    APIConfigurationError,
    APIConfigurationFailedError,
    APIKeyNotSetError,
    FileDeletingError,
    FileProcessingError,
    FileReadingError,
    FileWritingError,
    ProjectNotFoundError,
    ProjectProcessingError,
    WorkerError,
)


class Worker(Process):
    """プロジェクトをバックグラウンドで処理するワーカープロセス。"""

    def __init__(self, project_id: UUID, data_manager: DataManager) -> None:
        """
        Workerを初期化します。

        Args:
            project_id: 処理対象のプロジェクトID。
            data_manager: プロジェクトデータを管理するDataManagerインスタンス。
        """
        super().__init__()
        self.project_id = project_id
        self.data_manager = data_manager
        self.logger = logging.getLogger('aiman')

    @contextmanager
    def _handle_project_processing(self) -> Generator[Project, None, None]:
        """プロジェクト処理のコンテキストマネージャー。

        プロジェクトの取得、ステータス更新、エラーハンドリングを一元管理します。

        Yields:
            処理対象のProjectオブジェクト。

        Raises:
            ProjectProcessingError: プロジェクトが見つからない場合。
        """
        project = None
        try:
            project = self.data_manager.get_project(self.project_id)
            if not project:
                raise ProjectProcessingError(self.project_id)

            project.start_processing()
            self.data_manager._save_project(project)
            self.logger.info(f'Started processing project: {project.name}')
            yield project

        except Exception as e:
            self.logger.error(f'Error processing project {self.project_id}: {e}')
            if project:
                error_message = str(e)
                if isinstance(e, FileProcessingError):
                    error_message = f'ファイル処理エラー: {e}'
                elif isinstance(e, APIConfigurationError):
                    error_message = f'API設定エラー: {e}'
                project.fail({'error': error_message})
                self.data_manager._save_project(project)
            raise

    def run(self) -> None:
        """ワーカースレッドのメイン処理。プロジェクトの実行から完了までを担当します。"""
        self.logger.info(f'Worker started for project_id: {self.project_id}')

        try:
            with self._handle_project_processing() as project:
                if not config.GEMINI_API_KEY:
                    raise APIKeyNotSetError()

                genai.configure(api_key=config.GEMINI_API_KEY)
                model = genai.GenerativeModel(config.GEMINI_MODEL_NAME)

                self._execute_gemini_summarize(project, model)

        except (WorkerError, Exception) as e:
            self.logger.error(f'Worker error: {e}')

    def _process_xlsx(self, file_path: Path) -> tuple[str, list[Image.Image]]:
        """
        Excelファイルからテキストと画像を抽出します。
        画像はファイルから抽出し、テキスト内の画像があった場所には'[図:X]'という形式でマークが挿入されます。

        Args:
            file_path: 処理対象のExcelファイルのパス。

        Returns:
            抽出されたテキストと画像のリストのタプル。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            raise FileReadingError(str(e))

        full_text: list[str] = []
        images: list[Image.Image] = []
        image_count = 1

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f'---シート: {sheet_name}---\n')

            # 画像の位置を特定し、マーカーを挿入するための辞書
            image_markers: dict[tuple[int, int], str] = {}
            if sheet._images:  # type: ignore
                for img in sheet._images:  # type: ignore
                    # openpyxlの画像オブジェクトからPillowの画像オブジェクトに変換
                    try:
                        img_bytes = io.BytesIO(img._data())  # type: ignore
                        img_pil = Image.open(img_bytes)
                        images.append(img_pil)

                        # 画像のアンカーから位置を取得
                        row = img.anchor._from.row + 1  # type: ignore
                        col = img.anchor._from.col + 1  # type: ignore
                        marker = f'[図:{image_count}]'
                        image_markers[(row, col)] = marker
                        image_count += 1
                    except Exception as e:
                        self.logger.warning(f'Could not process an image in {file_path}: {e}')

            # セルの内容を読み込み、画像マーカーを挿入
            # セルの値とマーカーを結合するための辞書
            cell_values: dict[tuple[int, int], str] = {}
            for row_data in sheet.iter_rows():
                for cell in row_data:
                    if cell.value is not None:
                        cell_values[(cell.row, cell.column)] = str(cell.value)

            # マーカーの位置にマーカーを追加
            for (row, col), marker in image_markers.items():
                if (row, col) in cell_values:
                    cell_values[(row, col)] = marker + ' ' + cell_values[(row, col)]
                else:
                    cell_values[(row, col)] = marker

            # シートのテキストを再構築
            if cell_values:
                max_row = max(key[0] for key in cell_values.keys())
                max_col = max(key[1] for key in cell_values.keys())
                for r in range(1, max_row + 1):
                    row_text = []
                    for c in range(1, max_col + 1):
                        row_text.append(cell_values.get((r, c), ''))
                    full_text.append('\t'.join(row_text))

        return '\n'.join(full_text), images

    def _execute_gemini_summarize(self, project: Project, model: genai.GenerativeModel) -> None:
        """指定されたディレクトリ内のテキストファイルを要約し、結果をファイルに書き出します。

        Args:
            project: 処理対象のプロジェクトオブジェクト。
            model: 要約に使用するGeminiモデルのインスタンス。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
            APIConfigurationError: APIの設定や呼び出しに失敗した場合。
        """
        source_path = Path(project.source)
        output_file = source_path / 'gemini_results.md'
        processed_files = []

        if output_file.exists():
            try:
                output_file.unlink()
            except Exception as e:
                raise FileDeletingError(str(e))

        try:
            with open(output_file, 'w', encoding='utf-8') as f_out:
                f_out.write(f'# Gemini処理結果: {project.name}\n\n')

                # 複数の拡張子を検索
                files_to_process: list[Path] = []
                for ext in ['*.txt', '*.xlsx']:
                    files_to_process.extend(source_path.glob(ext))

                for file_path in files_to_process:
                    self.logger.info(f'Processing file: {file_path.name}')
                    try:
                        content: str = ''
                        images: list[Image.Image] = []
                        if file_path.suffix == '.txt':
                            try:
                                with open(file_path, encoding='utf-8') as f_in:
                                    content = f_in.read()
                            except Exception as e:
                                raise FileReadingError(str(e))
                        elif file_path.suffix == '.xlsx':
                            content, images = self._process_xlsx(file_path)

                        if not content.strip():
                            self.logger.warning(f'File {file_path.name} is empty. Skipping.')
                            continue

                        prompt: list[str | Image.Image] = [
                            f'以下の文章と図の内容を日本語で3行で要約し、各図の内容も簡単に解説してください。\n\n---\n{content}'
                        ]
                        # 画像があればプロンプトに追加
                        if images:
                            prompt.extend(images)

                        # prompt.jsonとして保存
                        # （画像はbase64エンコード、figure番号を付与）
                        prompt_json_path = file_path.parent / 'prompt.json'
                        prompt_for_json = []
                        # テキスト部分
                        for item in prompt:
                            if isinstance(item, str):
                                prompt_for_json.append({'type': 'text', 'data': item})
                        # 画像部分（figure番号を付与）
                        for idx, item in enumerate(prompt):
                            if isinstance(item, Image.Image):
                                buffered = io.BytesIO()
                                item.save(buffered, format='PNG')
                                img_b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                                prompt_for_json.append(
                                    {
                                        'type': 'image',
                                        'figure': str(idx),
                                        'data': img_b64,
                                    }
                                )
                        try:
                            with open(prompt_json_path, 'w', encoding='utf-8') as f_prompt:
                                json.dump(prompt_for_json, f_prompt, ensure_ascii=False, indent=2)
                        except Exception as e:
                            raise FileWritingError(str(e))

                        try:
                            response = model.generate_content(prompt)
                        except Exception as e:
                            raise APICallFailedError(e)

                        f_out.write(f'## ファイル: {file_path.name}\n\n')
                        f_out.write('### 要約結果\n\n')
                        f_out.write(response.text)
                        f_out.write('\n\n---\n\n')
                        processed_files.append(file_path.name)
                        time.sleep(config.API_RATE_LIMIT_DELAY)  # APIのレート制限を考慮

                    except FileProcessingError as e:
                        self.logger.error(f'Failed to process {file_path.name}: {e}')
                        f_out.write(f'## ファイル: {file_path.name}\n\n')
                        f_out.write(f'**エラー:** ファイル処理中にエラーが発生しました: {e}\n\n')
                        f_out.write('---\n\n')
                    except APIConfigurationError as e:
                        self.logger.error(f'API error while processing {file_path.name}: {e}')
                        f_out.write(f'## ファイル: {file_path.name}\n\n')
                        f_out.write(f'**エラー:** API呼び出し中にエラーが発生しました: {e}\n\n')
                        f_out.write('---\n\n')

        except Exception as e:
            raise FileWritingError(str(e))

        result = {
            'processed_files': processed_files,
            'message': f'Processing successful. Results saved to {output_file.name}',
        }

        self.data_manager.update_project_result(self.project_id, result)

    def _read_excel_file(self, file_path: Path) -> pd.DataFrame:
        """
        Excelファイルを読み込みます。

        Args:
            file_path: 読み込むExcelファイルのパス。

        Returns:
            読み込んだデータフレーム。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
        """
        try:
            return pd.read_excel(file_path)
        except Exception as e:
            raise FileReadingError(str(e))

    def _delete_output_file(self, file_path: Path) -> None:
        """
        既存の出力ファイルを削除します。

        Args:
            file_path: 削除するファイルのパス。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
        """
        try:
            file_path.unlink(missing_ok=True)
        except Exception as e:
            raise FileDeletingError(str(e))

    def _write_output_file(self, file_path: Path, content: str) -> None:
        """
        出力ファイルを書き込みます。

        Args:
            file_path: 書き込むファイルのパス。
            content: 書き込む内容。

        Raises:
            FileProcessingError: ファイルの処理に失敗した場合。
        """
        try:
            file_path.write_text(content, encoding='utf-8')
        except Exception as e:
            raise FileWritingError(str(e))

    def process_project(self) -> None:
        """
        プロジェクトを処理します。

        Raises:
            ProjectProcessingError: プロジェクトの処理中にエラーが発生した場合。
            FileProcessingError: ファイルの処理に失敗した場合。
            APIConfigurationError: APIの設定や呼び出しに失敗した場合。
        """
        try:
            project = self.data_manager.get_project(self.project_id)
            if project is None:
                raise ProjectNotFoundError(self.project_id)

            # Geminiモデルの初期化
            try:
                genai.configure(api_key=config.GEMINI_API_KEY)
                model = genai.GenerativeModel('gemini-pro-vision')
            except Exception as e:
                raise APIConfigurationFailedError(e)

            self._execute_gemini_summarize(project, model)

        except Exception as e:
            raise ProjectProcessingError(self.project_id) from e
