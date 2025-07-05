"""ExcelParserのテストケース。"""

import json
from collections.abc import Callable
from pathlib import Path

import openpyxl
import pytest
from PIL import Image

from app.errors import FileProcessingError, FileReadingError
from app.utils import ExcelParser
from app.utils.excel_parser import JsonDataDict


class TestExcelParser:
    """ExcelParserクラスのテストクラス。"""

    def test_初期化でインスタンスが正しく作成される(self) -> None:
        # Arrange
        # (準備なし)

        # Act
        parser = ExcelParser()

        # Assert
        assert parser._text_content == ''
        assert parser._images == []
        assert parser._parsed_data == {}

    def test_シンプルなExcelファイルを正しく解析できる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = 'Sheet1'
        sheet['A1'] = 'Hello'
        sheet['B1'] = 'World'
        sheet['A2'] = 'Test'
        sheet['B2'] = 'Data'
        workbook.save(test_file)
        parser = ExcelParser()

        # Act
        parser.parse(test_file)

        # Assert
        text = parser.get_text()
        assert '---シート: Sheet1---' in text
        assert 'Hello\tWorld' in text
        assert 'Test\tData' in text
        assert parser.get_images() == []

    def test_複数シートのExcelファイルを正しく解析できる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test_multi.xlsx'
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        assert sheet1 is not None
        sheet1.title = 'Sheet1'
        sheet1['A1'] = 'Sheet1 Data'
        sheet2 = workbook.create_sheet('Sheet2')
        sheet2['A1'] = 'Sheet2 Data'
        workbook.save(test_file)
        parser = ExcelParser()

        # Act
        parser.parse(test_file)

        # Assert
        text = parser.get_text()
        assert '---シート: Sheet1---' in text
        assert '---シート: Sheet2---' in text
        assert 'Sheet1 Data' in text
        assert 'Sheet2 Data' in text

    def test_空のExcelファイルを正しく解析できる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'empty.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(test_file)
        parser = ExcelParser()

        # Act
        parser.parse(test_file)

        # Assert
        text = parser.get_text()
        assert '---シート: Sheet---' in text  # デフォルトシート名
        assert parser.get_images() == []

    def test_存在しないファイルの解析時にエラーが発生する(self) -> None:
        # Arrange
        parser = ExcelParser()

        # Act & Assert
        with pytest.raises(FileReadingError):
            parser.parse(Path('nonexistent.xlsx'))

    def test_解析前のget_textは空文字列を返す(self) -> None:
        # Arrange
        parser = ExcelParser()

        # Act & Assert
        assert parser.get_text() == ''

    def test_解析前のget_imagesは空リストを返す(self) -> None:
        # Arrange
        parser = ExcelParser()

        # Act & Assert
        assert parser.get_images() == []

    def test_get_imagesは画像リストのコピーを返す(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(test_file)
        parser = ExcelParser()
        parser.parse(test_file)

        # Act
        images1 = parser.get_images()
        images2 = parser.get_images()

        # Assert
        assert images1 is not images2  # 異なるオブジェクト
        assert images1 == images2  # 内容は同じ

    def test_JSON変換が正しく行われる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = 'TestSheet'
        sheet['A1'] = 'Test Data'
        workbook.save(test_file)
        parser = ExcelParser()
        parser.parse(test_file)

        # Act
        json_data = parser.to_dict()

        # Assert
        assert json_data['file_path'] == str(test_file)
        assert json_data['total_images'] == 0
        assert 'TestSheet' in json_data['sheets']
        assert 'Test Data' in json_data['text_content']
        assert json_data['images'] == []

    def test_解析前のJSON変換は空のデータを返す(self) -> None:
        # Arrange
        parser = ExcelParser()

        # Act
        json_data = parser.to_dict()

        # Assert
        assert json_data['images'] == []
        assert json_data['text_content'] == ''

    def test_JSONファイルの出力が正しく行われる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet['A1'] = 'Export Test'
        workbook.save(test_file)
        parser = ExcelParser()
        parser.parse(test_file)
        output_file = tmp_path / 'output.json'

        # Act
        parser.export_json(output_file)

        # Assert
        assert output_file.exists()
        with open(output_file, encoding='utf-8') as f:
            data = json.load(f)
        assert 'Export Test' in data['text_content']
        assert data['total_images'] == 0

    def test_無効なパスへのJSON出力でエラーが発生する(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'test.xlsx'
        workbook = openpyxl.Workbook()
        workbook.save(test_file)
        parser = ExcelParser()
        parser.parse(test_file)
        invalid_path = tmp_path / 'nonexistent_dir' / 'output.json'

        # Act & Assert
        with pytest.raises(FileProcessingError):
            parser.export_json(invalid_path)

    def _extract_cell_values_from_sheet(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
    ) -> dict[tuple[int, int], str]:
        """シートからセル値を抽出する"""
        cell_values = {}
        for row_data in sheet.iter_rows():
            for cell in row_data:
                if cell.value is not None:
                    cell_values[(cell.row, cell.column)] = str(cell.value)
        return cell_values

    def _build_text_from_cell_values(
        self,
        cell_values: dict[tuple[int, int], str],
        full_text: list[str],
    ) -> None:
        """セル値からテキストを構築する"""
        if cell_values:
            keys = list(cell_values.keys())
            max_row = max(key[0] for key in keys)
            max_col = max(key[1] for key in keys)
            for r in range(1, max_row + 1):
                row_text = []
                for c in range(1, max_col + 1):
                    row_text.append(cell_values.get((r, c), ''))
                full_text.append('\t'.join(row_text))

    def _create_mock_process_with_image(
        self,
        test_image: Image.Image,
    ) -> Callable[[ExcelParser, Path], None]:
        """画像を含むExcelファイルの処理をシミュレートするモック関数を作成する"""

        def mock_process_with_image(parser: ExcelParser, file_path: Path) -> None:
            workbook = openpyxl.load_workbook(file_path)
            full_text = ['---シート: Sheet---\n']
            sheet = workbook.active
            assert sheet is not None

            cell_values = self._extract_cell_values_from_sheet(sheet)
            cell_values[(1, 1)] = '[図:1] ' + cell_values.get((1, 1), '')
            self._build_text_from_cell_values(cell_values, full_text)

            parser._text_content = '\n'.join(full_text)
            parser._images = [test_image]
            serializable_cell_values = {
                f'{row},{col}': value for (row, col), value in cell_values.items()
            }
            parser._parsed_data = {
                'file_path': str(file_path),
                'total_images': 1,
                'sheets': {
                    'Sheet': {
                        'text_lines': full_text[1:],
                        'images': [{'figure_number': 1, 'row': 1, 'col': 1, 'marker': '[図:1]'}],
                        'cell_values': serializable_cell_values,
                    },
                },
            }

        return mock_process_with_image

    def _assert_image_simulation_results(
        self,
        text: str,
        images: list[Image.Image],
        json_data: JsonDataDict,
    ) -> None:
        """画像シミュレーション結果のアサーションを実行する"""
        assert '[図:1]' in text
        assert len(images) == 1
        assert isinstance(images[0], Image.Image)
        assert json_data['total_images'] == 1
        assert len(json_data['images']) == 1
        assert json_data['images'][0]['figure_number'] == 1

    def _assert_complex_excel_results(self, text: str, json_data: JsonDataDict) -> None:
        """複雑なExcel構造の解析結果を検証するヘルパーメソッド"""
        assert 'Header1\tHeader2\tHeader3' in text
        assert 'Data1\tData2\tData3' in text
        assert 'Separated Data' in text
        assert 'ComplexSheet' in json_data['sheets']
        sheet_data = json_data['sheets']['ComplexSheet']
        assert '1,1' in sheet_data['cell_values']
        assert sheet_data['cell_values']['1,1'] == 'Header1'

    def test_画像を含むExcelファイルのシミュレーション処理が正しく行われる(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        # Arrange
        test_image = Image.new('RGB', (100, 100), color='red')
        test_file = tmp_path / 'test_with_image.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet['A1'] = 'Text with image'

        mock_function = self._create_mock_process_with_image(test_image)
        monkeypatch.setattr(ExcelParser, 'parse', mock_function)
        workbook.save(test_file)
        parser = ExcelParser()

        # Act
        parser.parse(test_file)
        text = parser.get_text()
        images = parser.get_images()
        json_data = parser.to_dict()

        # Assert
        self._assert_image_simulation_results(text, images, json_data)

    def _create_complex_excel_file(self, test_file: Path) -> None:
        """複雑なExcelファイルを作成するヘルパーメソッド"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = 'ComplexSheet'
        sheet['A1'] = 'Header1'
        sheet['B1'] = 'Header2'
        sheet['C1'] = 'Header3'
        sheet['A2'] = 'Data1'
        sheet['B2'] = 'Data2'
        sheet['C2'] = 'Data3'
        sheet['A4'] = 'Separated Data'  # 空行を挨んだデータ
        workbook.save(test_file)

    def test_複雑なExcel構造の解析が正しく行われる(self, tmp_path: Path) -> None:
        # Arrange
        test_file = tmp_path / 'complex.xlsx'
        self._create_complex_excel_file(test_file)
        parser = ExcelParser()

        # Act
        parser.parse(test_file)
        text = parser.get_text()
        json_data = parser.to_dict()

        # Assert
        self._assert_complex_excel_results(text, json_data)
